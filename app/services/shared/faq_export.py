"""FAQ 导出服务

提供 FAQ 数据的导出功能，支持 CSV、JSON、Excel 格式。
对齐 WeKnora 的 FAQ 导出 API 实现。
"""

import csv
import json
from datetime import UTC, datetime
from enum import Enum
from io import BytesIO, StringIO
from typing import Annotated, Any

from fastapi import Depends
from sqlalchemy.ext.asyncio import AsyncSession

from app.infra.database import get_session
from app.models.faq import FAQ, FAQCategory, FAQStatus
from app.observability.logging import get_logger
from app.repositories.faq import FAQRepository

logger = get_logger(__name__)


class ExportFormat(str, Enum):
    """导出格式"""

    CSV = "csv"
    JSON = "json"
    EXCEL = "excel"


class FAQExporter:
    """FAQ 导出器

    提供多种格式的 FAQ 数据导出功能。
    """

    # CSV 导出列定义（对齐 WeKnora）
    CSV_COLUMNS = [
        "id",
        "question",
        "answer",
        "category",
        "tags",
        "priority",
        "locale",
        "status",
        "slug",
        "view_count",
        "helpful_count",
        "not_helpful_count",
        "created_at",
        "updated_at",
        "published_at",
    ]

    def __init__(self, session: AsyncSession) -> None:
        """初始化导出器

        Args:
            session: 数据库会话
        """
        self.session = session
        self._repository: FAQRepository | None = None

    @property
    def repository(self) -> FAQRepository:
        """获取仓储（延迟初始化）"""
        if self._repository is None:
            self._repository = FAQRepository(self.session)
        return self._repository

    async def _fetch_all_faqs(
        self,
        *,
        status: FAQStatus | None = None,
        category: FAQCategory | None = None,
        locale: str | None = None,
        tenant_id: int | None = None,
    ) -> list[FAQ]:
        """获取所有 FAQ 数据

        Args:
            status: 状态筛选
            category: 分类筛选
            locale: 语言筛选
            tenant_id: 租户 ID

        Returns:
            FAQ 列表
        """
        from sqlalchemy import or_, select

        statement = select(FAQ)

        if status:
            statement = statement.where(FAQ.status == status)

        if category:
            statement = statement.where(FAQ.category == category)

        if locale:
            statement = statement.where(FAQ.locale == locale)

        if tenant_id is not None:
            statement = statement.where(
                or_(FAQ.tenant_id == tenant_id, FAQ.tenant_id.is_(None))
            )

        statement = statement.order_by(FAQ.category, FAQ.priority, FAQ.id)

        result = await self.session.execute(statement)
        return list(result.scalars().all())

    def _faq_to_dict(self, faq: FAQ) -> dict[str, Any]:
        """将 FAQ 转换为字典

        Args:
            faq: FAQ 模型

        Returns:
            字典形式的 FAQ 数据
        """
        return {
            "id": faq.id or 0,
            "question": faq.question,
            "answer": faq.answer,
            "category": faq.category.value,
            "tags": json.dumps(faq.tags or [], ensure_ascii=False),
            "priority": faq.priority,
            "locale": faq.locale,
            "status": faq.status.value,
            "slug": faq.slug or "",
            "view_count": faq.view_count,
            "helpful_count": faq.helpful_count,
            "not_helpful_count": faq.not_helpful_count,
            "created_at": (
                faq.created_at.isoformat() if faq.created_at else ""
            ),
            "updated_at": (
                faq.updated_at.isoformat() if faq.updated_at else ""
            ),
            "published_at": (
                faq.published_at.isoformat() if faq.published_at else ""
            ),
        }

    async def export_to_csv(
        self,
        *,
        status: FAQStatus | None = None,
        category: FAQCategory | None = None,
        locale: str | None = None,
        tenant_id: int | None = None,
    ) -> tuple[str, str]:
        """导出为 CSV 格式

        Args:
            status: 状态筛选
            category: 分类筛选
            locale: 语言筛选
            tenant_id: 租户 ID

        Returns:
            (文件名, CSV 内容)
        """
        faqs = await self._fetch_all_faqs(
            status=status,
            category=category,
            locale=locale,
            tenant_id=tenant_id,
        )

        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=self.CSV_COLUMNS)
        writer.writeheader()

        for faq in faqs:
            row = self._faq_to_dict(faq)
            writer.writerow(row)

        content = output.getvalue()

        filename = self._generate_filename("csv")

        logger.info(
            "faq_exported_csv",
            count=len(faqs),
            filename=filename,
            filters={
                "status": status.value if status else None,
                "category": category.value if category else None,
                "locale": locale,
            },
        )

        return filename, content

    async def export_to_json(
        self,
        *,
        status: FAQStatus | None = None,
        category: FAQCategory | None = None,
        locale: str | None = None,
        tenant_id: int | None = None,
    ) -> tuple[str, str]:
        """导出为 JSON 格式

        Args:
            status: 状态筛选
            category: 分类筛选
            locale: 语言筛选
            tenant_id: 租户 ID

        Returns:
            (文件名, JSON 内容)
        """
        faqs = await self._fetch_all_faqs(
            status=status,
            category=category,
            locale=locale,
            tenant_id=tenant_id,
        )

        data = {
            "exported_at": datetime.now(UTC).isoformat(),
            "total": len(faqs),
            "filters": {
                "status": status.value if status else None,
                "category": category.value if category else None,
                "locale": locale,
            },
            "items": [
                {
                    "id": faq.id or 0,
                    "question": faq.question,
                    "answer": faq.answer,
                    "category": faq.category.value,
                    "tags": faq.tags or [],
                    "priority": faq.priority,
                    "locale": faq.locale,
                    "status": faq.status.value,
                    "slug": faq.slug,
                    "view_count": faq.view_count,
                    "helpful_count": faq.helpful_count,
                    "not_helpful_count": faq.not_helpful_count,
                    "created_at": (
                        faq.created_at.isoformat() if faq.created_at else None
                    ),
                    "updated_at": (
                        faq.updated_at.isoformat() if faq.updated_at else None
                    ),
                    "published_at": (
                        faq.published_at.isoformat() if faq.published_at else None
                    ),
                }
                for faq in faqs
            ],
        }

        content = json.dumps(data, ensure_ascii=False, indent=2)

        filename = self._generate_filename("json")

        logger.info(
            "faq_exported_json",
            count=len(faqs),
            filename=filename,
        )

        return filename, content

    async def export_to_excel(
        self,
        *,
        status: FAQStatus | None = None,
        category: FAQCategory | None = None,
        locale: str | None = None,
        tenant_id: int | None = None,
    ) -> tuple[str, bytes]:
        """导出为 Excel 格式

        Args:
            status: 状态筛选
            category: 分类筛选
            locale: 语言筛选
            tenant_id: 租户 ID

        Returns:
            (文件名, Excel 二进制内容)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            filename, csv_content = await self.export_to_csv(
                status=status,
                category=category,
                locale=locale,
                tenant_id=tenant_id,
            )
            return filename, csv_content.encode("utf-8")

        faqs = await self._fetch_all_faqs(
            status=status,
            category=category,
            locale=locale,
            tenant_id=tenant_id,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "FAQ"

        # 设置标题行
        header_font = Font(bold=True)
        for col_idx, column_name in enumerate(self.CSV_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 填充数据
        for row_idx, faq in enumerate(faqs, start=2):
            row_data = self._faq_to_dict(faq)
            for col_idx, column_name in enumerate(self.CSV_COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data[column_name]
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        cell_value = str(cell.value)
                        # 计算中文字符（占 2 个宽度）
                        length = sum(
                            2 if ord(char) > 127 else 1
                            for char in cell_value[:50]  # 限制检查长度
                        )
                        max_length = max(max_length, length)
                except ValueError:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度 50
            ws.column_dimensions[column].width = adjusted_width

        # 保存到内存

        output = BytesIO()
        wb.save(output)
        content = output.getvalue()

        filename = self._generate_filename("xlsx")

        logger.info(
            "faq_exported_excel",
            count=len(faqs),
            filename=filename,
        )

        return filename, content

    def _generate_filename(self, extension: str) -> str:
        """生成导出文件名

        Args:
            extension: 文件扩展名

        Returns:
            文件名
        """
        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        return f"faq_export_{timestamp}.{extension}"


# ============== 依赖注入工厂 ==============


def get_faq_exporter(
    session: Annotated[AsyncSession, Depends(get_session)],
) -> FAQExporter:
    """创建 FAQ 导出器实例

    Args:
        session: 数据库会话

    Returns:
        FAQExporter 实例
    """
    return FAQExporter(session)
